#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Regenera index.html del tablero de Impacto en Ventas de Ensifera.

Descarga la hoja publica «Conglomerado Gastos Publicitarios 2026», la parsea,
la combina con meta.json (instantanea de Meta Ads Ensifera COP) y la inyecta
en template.html -> index.html. Disenado para correr desatendido (rutina nube).

Uso:  python build.py
Req:  pip install openpyxl
"""
import io, os, re, json, sys, datetime, urllib.request
from openpyxl import load_workbook

FILE_ID = "1sxBiAsb0sUOLqYNr3H7l-VhN6DpN-Boh"
URL = "https://docs.google.com/spreadsheets/d/%s/export?format=xlsx" % FILE_ID
USDCOP = 3600
HERE = os.path.dirname(os.path.abspath(__file__))
MONTHNUM = {"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
            "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}

def s(x):
    return "" if x is None else str(x).strip()

def is_num(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool)

def cell(row, i):
    return row[i] if (row is not None and 0 <= i < len(row)) else None

def cint(x):
    if x is None: return None
    if is_num(x): return float(x)
    st = str(x)
    if "#" in st: return None
    d = re.sub(r"[^0-9]", "", st)
    return float(d) if d else None

def cusd(x):
    if x is None: return None
    if is_num(x): return float(x)
    st = re.sub(r"[\$\s]", "", str(x).strip())
    if st in ("", "-") or "#" in st: return None
    if re.match(r"^[0-9]+,[0-9]{1,2}$", st): st = st.replace(",", ".")
    else: st = st.replace(",", "")
    try: return float(st)
    except ValueError: return None

def cpct(x):
    if x is None: return None
    if is_num(x): return float(x) * 100.0
    st = str(x)
    if "#" in st: return None
    st = re.sub(r"[%\s]", "", st).replace(",", ".")
    try: return float(st)
    except ValueError: return None

def croas(x):
    if x is None: return None
    if is_num(x): return float(x)
    st = str(x)
    if "#" in st: return None
    st = re.sub(r"[\$\s]", "", st).replace(",", ".")
    if st in ("", "-"): return None
    try: return float(st)
    except ValueError: return None

def pdate(x):
    if x is None: return None
    if isinstance(x, (datetime.datetime, datetime.date)):
        return x.strftime("%Y-%m-%d")
    if is_num(x):
        n = float(x)
        if 20000 < n < 90000:
            return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=n)).strftime("%Y-%m-%d")
        return None
    st = str(x).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", st)
    if m: return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", st)
    if m: return "%04d-%02d-%02d" % (int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None

def rnd(v):
    return None if v is None else int(round(v))

def parse_sheet(ws):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    n = len(rows)
    ncol = ws.max_column or 0
    markers = []
    for i in range(n):
        m = re.match(r"^GASTOS PUBLICITARIOS\s+(\d+)\s+([A-Za-zÀ-ɏ]+)", s(cell(rows[i], 0)))
        if m: markers.append((i, int(m.group(1)), m.group(2).upper()))
    out = []
    for mi in range(len(markers)):
        start = markers[mi][0]
        end = markers[mi + 1][0] if mi + 1 < len(markers) else n
        campaigns, sedes_conv, agents, products, sales_by_sede, summ = [], {}, {}, {}, {}, {}
        date = None; gasto_usd = gasto_cop = None
        impr = reach = res = 0.0
        r = start
        while r < end:
            row = rows[r]
            c0 = s(cell(row, 0))
            # campaign/ad table (header contains "Importe gastado (USD|COP)")
            spend_col = -1; curr = None
            for c in range(ncol):
                h = s(cell(row, c))
                if "Importe gastado" in h:
                    spend_col = c; curr = "USD" if "USD" in h else "COP"; break
            if spend_col >= 0:
                name_col = impr_col = reach_col = res_col = -1
                for c in range(ncol):
                    h = s(cell(row, c))
                    if h.startswith("Nombre"): name_col = c
                    elif h == "Impresiones": impr_col = c
                    elif h == "Alcance": reach_col = c
                    elif re.match(r"^(Resultados|Mensajes totales|Contactos de mensajes)$", h): res_col = c
                if name_col < 0: name_col = 0
                j = r + 1
                while j < end:
                    r2 = rows[j]; gt = False
                    for c in range(ncol):
                        if s(cell(r2, c)) == "GASTO TOTAL":
                            gt = True
                            if curr == "USD": gasto_usd = cusd(cell(r2, c + 1))
                            else: gasto_cop = cint(cell(r2, c + 1))
                            break
                    if gt: break
                    nm = s(cell(r2, name_col))
                    if nm == "" or re.match(r"^(Nombre|Inicio)", nm):
                        j += 1; continue
                    if curr == "USD":
                        u = cusd(cell(r2, spend_col)); sp = rnd(u * 3600) if u is not None else None
                    else:
                        sp = cint(cell(r2, spend_col))
                    im = cint(cell(r2, impr_col)) if impr_col >= 0 else None
                    rc = cint(cell(r2, reach_col)) if reach_col >= 0 else None
                    rs = cint(cell(r2, res_col)) if res_col >= 0 else None
                    campaigns.append({"name": nm, "spendCOP": sp, "impressions": im, "reach": rc, "results": rs})
                    if im: impr += im
                    if rc: reach += rc
                    if rs: res += rs
                    j += 1
                r = j; continue
            # CONVERSACIONES (agentes + sedes)
            if re.match(r"^CONVERSACIONES", c0):
                a_col = h_row = -1
                for j in range(r + 1, min(r + 4, end)):
                    for c in range(ncol):
                        if s(cell(rows[j], c)) == "Agente": a_col = c; h_row = j; break
                    if a_col >= 0: break
                if a_col >= 0:
                    j = h_row + 1
                    while j < end:
                        a = s(cell(rows[j], a_col))
                        if a == "" or re.match(r"^(MENSAJES|PRODUCTOS|CONVERSACIONES|Ventas|Valor|NOMBRE)", a): break
                        cnt = cint(cell(rows[j], a_col + 1))
                        if cnt is not None:
                            if re.match(r"(?i)^Ensifera\s", a): sedes_conv[a] = cnt
                            else: agents[a] = cnt
                        j += 1
                r += 1; continue
            # PRODUCTOS
            if c0 == "PRODUCTOS":
                p_col = h_row = -1
                for j in range(r + 1, min(r + 4, end)):
                    for c in range(ncol):
                        if s(cell(rows[j], c)) == "Producto": p_col = c; h_row = j; break
                    if p_col >= 0: break
                if p_col >= 0:
                    j = h_row + 1
                    while j < end:
                        p = s(cell(rows[j], p_col))
                        if p == "" or re.match(r"^(MENSAJES|CONVERSACIONES|Ventas|Valor|NOMBRE)", p): break
                        cnt = cint(cell(rows[j], p_col + 1))
                        if cnt is not None: products[p] = cnt
                        j += 1
                r += 1; continue
            # Ventas Totales <Sede>
            if re.match(r"^Ventas Totales\s+(Palmira|Cali|Medell|Bogot)", c0):
                sede = re.sub(r"^Ventas Totales\s+", "", c0)
                v = cint(cell(row, 1))
                sales_by_sede[sede] = rnd(v) if v is not None else 0
                r += 1; continue
            # Ventas Totales standalone (fila agregado antes del bloque NOMBRE)
            if c0 == "Ventas Totales":
                raw = cell(row, 1)
                v = cint(raw)
                print(f"  [DBG] standalone VT row={r+1} raw={raw!r} cint={v}")
                if v is not None:
                    summ["salesCOP"] = rnd(v)
                r += 1; continue
            # trailing NOMBRE summary
            if c0 == "NOMBRE":
                dd = pdate(cell(row, 1))
                if dd: date = dd
                for j in range(r + 1, min(r + 11, end)):
                    lbl = s(cell(rows[j], 0)); val = cell(rows[j], 1)
                    if re.search(r"Inversi.n USD", lbl): summ["spendUSD"] = cusd(val)
                    elif re.search(r"Inversion PC", lbl): summ["spendCOP"] = cint(val)
                    elif lbl == "Conversaciones": summ["conversations"] = cint(val)
                    elif re.search(r"Costo Conversaci", lbl): summ["costConvCOP"] = cint(val)
                    elif lbl == "Ventas Totales":
                        v = cint(val)
                        if v is not None:
                            summ["salesCOP"] = v
                    elif re.search(r"Inversion Venta", lbl): summ["invSalePct"] = cpct(val)
                    elif lbl == "ROAS": summ["roas"] = croas(val)
                    elif re.search(r"Ticket Conversi", lbl): summ["ticketConvPct"] = cpct(val)
                r += 1; continue
            r += 1

        if not date:
            date = "2026-%02d-%02d" % (MONTHNUM.get(markers[mi][2], 1), markers[mi][1])
        spend_cop = summ.get("spendCOP")
        if spend_cop is None:
            spend_cop = gasto_cop if gasto_cop is not None else (rnd(gasto_usd * 3600) if gasto_usd is not None else None)
        out.append({
            "date": date, "month": ws.title,
            "spendCOP": rnd(spend_cop),
            "spendUSD": summ.get("spendUSD"),
            "impressions": int(impr), "reach": int(reach), "resultsAds": int(res),
            "conversations": rnd(summ.get("conversations")),
            "costConvCOP": rnd(summ.get("costConvCOP")),
            "salesCOP": rnd(summ.get("salesCOP")),
            "roas": summ.get("roas"),
            "invSalePct": summ.get("invSalePct"),
            "ticketConvPct": summ.get("ticketConvPct"),
            "salesBySede": sales_by_sede, "sedesConv": sedes_conv,
            "agents": agents, "products": products, "campaigns": campaigns,
        })
    return out

def main():
    sys.stderr.write("Descargando hoja publica...\n")
    data = urllib.request.urlopen(URL, timeout=90).read()
    if data[:2] != b"PK":
        sys.exit("ERROR: la descarga no es un xlsx. La hoja debe estar compartida como 'Cualquiera con el enlace (Lector)'.")
    wb = load_workbook(io.BytesIO(data), data_only=True)
    all_days = []
    for ws in wb.worksheets:
        all_days.extend(parse_sheet(ws))
    by_date = {}
    for d in sorted(all_days, key=lambda d: d["date"]):
        by_date[d["date"]] = d

    # Suplementar con meta_daily.json: agrega días de Meta Ads que no están en la hoja
    meta_daily_path = os.path.join(HERE, "meta_daily.json")
    if os.path.exists(meta_daily_path):
        with open(meta_daily_path, encoding="utf-8") as mf:
            meta_data = json.load(mf)
        n_added = 0
        for md in meta_data.get("days", []):
            if md["date"] not in by_date:
                mo_num = int(md["date"][5:7])
                mo_name = list(MONTHNUM.keys())[mo_num - 1]
                by_date[md["date"]] = {
                    "date": md["date"], "month": mo_name.capitalize(),
                    "spendCOP": md.get("spendCOP"), "spendUSD": None,
                    "impressions": md.get("impressions", 0), "reach": 0,
                    "resultsAds": md.get("resultsAds", 0), "conversations": None,
                    "costConvCOP": None, "salesCOP": None, "roas": None,
                    "invSalePct": None, "ticketConvPct": None,
                    "salesBySede": {}, "sedesConv": {}, "agents": {}, "products": {},
                    "campaigns": [{"name": c["name"], "spendCOP": c.get("spendCOP"),
                                   "impressions": c.get("impressions"), "reach": None,
                                   "results": c.get("results")}
                                  for c in md.get("campaigns", [])],
                }
                n_added += 1
        if n_added:
            sys.stderr.write(f"  + {n_added} dias adicionales de meta_daily.json\n")

    days = sorted(by_date.values(), key=lambda d: d["date"])
    if not days:
        sys.exit("ERROR: no se encontraron dias en la hoja.")

    with open(os.path.join(HERE, "meta.json"), encoding="utf-8") as f:
        meta = json.load(f)
    combined = {"usdCop": USDCOP, "days": days, "meta": {
        "account": meta.get("account"), "accountId": meta.get("accountId"),
        "currency": meta.get("currency"), "period": meta.get("period"),
        "campaignTotals": meta.get("campaignTotals", []), "daily": meta.get("daily", [])}}
    data_json = json.dumps(combined, ensure_ascii=False, separators=(",", ":"))

    with open(os.path.join(HERE, "template.html"), encoding="utf-8") as f:
        tpl = f.read()
    if "/*__DATA__*/ {}" not in tpl:
        sys.exit("ERROR: falta el placeholder /*__DATA__*/ {} en template.html")
    with open(os.path.join(HERE, "index.html"), "w", encoding="utf-8") as f:
        f.write(tpl.replace("/*__DATA__*/ {}", data_json))

    t_spend = sum(d["spendCOP"] or 0 for d in days)
    t_sales = sum(d["salesCOP"] or 0 for d in days)
    sys.stderr.write("OK: %d dias (%s -> %s)  inversion=%d COP  ventas=%d COP\n" %
                     (len(days), days[0]["date"], days[-1]["date"], t_spend, t_sales))

if __name__ == "__main__":
    main()
